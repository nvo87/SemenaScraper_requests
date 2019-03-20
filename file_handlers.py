import os
import csv
import json
import string
import urllib.request
from urllib.parse import urlparse
import concurrent.futures

import openpyxl


class FileHandler:
    """
        To init handler in main program:
            from file_handlers import FileHandler
            class Scraper:
                def __init__(self):
                    self.file = FileHandler

            s = Scraper()
            file = s.file('2.txt').handler
            urls_list = file.read_data_to_list()
    """
    # def __init__(self, filename):
    #     self.filename = filename
    #     self.handler = self._init_handler_by_extension()
    #
    # def _init_handler_by_extension(self):
    #     name, extension = os.path.splitext(self.filename)
    #     return {
    #         '.csv': CSVHandler,
    #         '.txt': TxtHandler(self.filename),
    #         '.xls': ExcelHandler(self.filename),
    #         '.xlsx': ExcelHandler(self.filename),
    #         '.json': JSONHandler(self.filename),
    #     }.get(extension, None)


    def open_clear_or_create(self):
        with open(self.filename, 'w') as f:
            pass

    def open_to_append_or_create(self):
        with open(self.filename, 'a') as f:
            pass


class CSVHandler(FileHandler):
    def __init__(self, filename):
        self.filename = filename

    def add_row_from_list(self, row_data: list):
        with open(self.filename, 'a') as f:
            w = csv.writer(f)
            w.writerow(
                [''.join(char for char in el if char in string.printable) for el in row_data])

    def add_row_from_dict(self, row_data: dict, fieldnames=[]):
        with open(self.filename, 'a') as f:
            fieldnames = fieldnames or row_data.keys()
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writerow(row_data)

    def write_list_of_dict(self, list_of_dict, with_header=False):
        keys = list_of_dict[0].keys()
        with open(self.filename, 'w') as f:
            dict_writer = csv.DictWriter(f, keys)
            if with_header:
                dict_writer.writeheader()
            dict_writer.writerows(list_of_dict)

    def read_to_list_of_dicts(self):
        with open(self.filename, 'r') as f:
            list_of_dicts = [{k: v for k, v in row.items()}
                             for row in csv.DictReader(f, skipinitialspace=True)]
            return list_of_dicts


class TxtHandler(FileHandler):
    def __init__(self, filename):
        self.filename = filename

    def print_to_file(self, data, mode='w'):
        with open(self.filename, mode) as f:
            if type(data) in (str, float, int):
                print(data, file=f, sep='\n')
            else:
                print(*data, file=f, sep='\n')

    def read_data_to_list(self):
        with open(self.filename, 'r') as f:
            return f.read().splitlines()


class ExcelHandler(FileHandler):
    """
    wb - WorkBook
    ws - WorkSheet
    """
    def __init__(self, filename):
        self.wb = None
        self.active_ws = None
        self.filename = filename

    def open_workbook(self, read_only=False):
        self.wb = openpyxl.load_workbook(self.filename, read_only=read_only)

    def open_worksheet(self, ws_name=''):
        if ws_name:
            self.active_ws = self.wb[ws_name]
        else:
            self.active_ws = self.wb.active

    def create_workbook(self, write_only=False):
        self.wb = openpyxl.Workbook(write_only=write_only)

    def create_worksheet(self, ws_name, ws_position: int = None):
        self.active_ws = self.wb.create_sheet(ws_name, ws_position)

    def append_row_to_sheet(self, row: list):
        self.active_ws.append(row)

    def save_workbook(self):
        self.wb.save(self.filename)


class JSONHandler(FileHandler):
    def __init__(self, filename):
        self.filename = filename

    def write_json_to_file(self, list_of_dict):
        with open(self.filename, 'w', encoding='utf-8') as f:
            json.dump(list_of_dict, f, sort_keys=True, indent=4, ensure_ascii=False)

    def read_json_from_file(self):
        with open(self.filename, encoding='utf-8') as f:
            return json.load(f)


class FileDownloader:
    def download_files_from_urls_list(self, urls_list, result_dir=''):
        if not os.path.exists(result_dir):
            os.makedirs(result_dir)
        # We can use a with statement to ensure threads are cleaned up promptly
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            # Start the load operations and mark each future with its URL
            future_to_url = {executor.submit(self._load_url, url, 45): url for url in urls_list}
            for future, file_url in zip(concurrent.futures.as_completed(future_to_url), urls_list):
                url = future_to_url[future]
                file_name = os.path.basename(urlparse(file_url).path)
                file_path = os.path.join(result_dir, file_name)
                try:
                    data = future.result()
                    with open(file_path, 'wb') as handler:
                        handler.write(data)
                except Exception as exc:
                    print('%r generated an exception: %s' % (url, exc))
                else:
                    print('%r page is %d bytes' % (url, len(data)))

    # Retrieve a single page and report the URL and contents
    def _load_url(self, url, timeout):
        with urllib.request.urlopen(url, timeout=timeout) as conn:
            return conn.read()


def main():
    # csv = CSVHandler(filename)
    # data = csv.read_to_list_of_dicts()
    # urls = [item['img_url'] for item in data]
    # saver = FileDownloader()
    # saver.download_files_from_urls_list(urls, './result')
    pass

if __name__ == '__main__':
    main()
